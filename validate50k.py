import pandas as pd
from openpyxl import Workbook
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Global variables to store the model and tokenizer
global_tokenizers = {}
global_models = {}

def translate_text(src_lang, tar_lang, original_txt):
    model_name = f"Eugenememe/netflix-{src_lang}-{tar_lang}-50k"

    # Check if the model and tokenizer are already loaded
    if model_name not in global_models or model_name not in global_tokenizers:
        global_tokenizers[model_name] = AutoTokenizer.from_pretrained(model_name)
        global_models[model_name] = AutoModelForSeq2SeqLM.from_pretrained(model_name)

    tokenizer = global_tokenizers[model_name]
    model = global_models[model_name]

    translated_txt = []
    for text in original_txt:
        translated = model.generate(**tokenizer([text], return_tensors="pt", padding=True))
        translation = tokenizer.decode(translated[0], skip_special_tokens=True)
        translated_txt.append(translation)
    
    return translated_txt

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def auto_adjust_columns(sheet, min_width=10, max_width=50, estimated_char_per_line=50):
    for column_cells in sheet.columns:
        max_text_length = max(len(str(cell.value)) for cell in column_cells)
        column_width = min(max(max_text_length, min_width) + 2, max_width)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = column_width

        # Set text wrapping, font, and size
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True)
            cell.font = Font(name="Arial", size=12)

            # Calculate required row height
            text_lines = len(str(cell.value)) // estimated_char_per_line + 1
            required_row_height = text_lines * 15  # 15 is an estimated height per line

            current_row_height = sheet.row_dimensions[cell.row].height
            if not current_row_height or current_row_height < required_row_height:
                sheet.row_dimensions[cell.row].height = required_row_height

    # Header formatting
    for cell in sheet[1]:
        cell.font = Font(name="Arial", size=12, bold=True)
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

def write_to_excel(original_file_path, language_pairs, excel_file_path):
    workbook = Workbook()
    original_text = []

    with open(original_file_path, 'r', encoding='utf-8') as file:
        original_text = [line.strip() for line in file]

    for src_lang, tar_lang in language_pairs:
        translated_txt = translate_text(src_lang, tar_lang, original_text)
        sheet = workbook.create_sheet(title=f"{src_lang}-{tar_lang}")
        sheet.append(["Original Text", f"{src_lang} to {tar_lang}"])
        for orig, trans in zip(original_text, translated_txt):
            sheet.append([orig, trans])
        auto_adjust_columns(sheet)

    # Delete the default sheet created by openpyxl
    del workbook['Sheet']
    workbook.save(excel_file_path)

# Main Process
# original_file_path = "original_english_text.txt"
original_file_path = "new_ver_original_english_text.txt"
# language_pairs = [("en", "zh"), ("en", "fr"), ("en", "de")]
language_pairs = [("en", "es")]
excel_file_path = "en_es_50k_translations.xlsx"

write_to_excel(original_file_path, language_pairs, excel_file_path)